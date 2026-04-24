"""Build a styled multi-sheet Excel report using openpyxl."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# --- Style constants ---
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")       # light blue
KPI_FILL = PatternFill("solid", fgColor="2E75B6")       # medium blue
KPI_LABEL_FILL = PatternFill("solid", fgColor="1F3864")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
KPI_VALUE_FONT = Font(bold=True, color="FFFFFF", size=16)
KPI_LABEL_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1F3864", size=14)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply_border(cell):
    cell.border = THIN_BORDER


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Write a DataFrame to a worksheet with styled header and alternating rows.
    Returns the last row written."""
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(cell)

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = ALT_FILL if (row_idx - start_row) % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _apply_border(cell)

    # Auto column widths
    for col_idx, header in enumerate(headers, start=1):
        col_values = [str(header)] + [str(v) for v in df.iloc[:, col_idx - 1]]
        max_len = max(len(v) for v in col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    return start_row + len(df)


def _add_bar_chart(ws, data_ws_title: str, wb: Workbook, data_start_row: int, num_rows: int, num_cols: int, chart_anchor: str, title: str):
    """Add a bar chart to ws, reading data from data_ws_title."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Category"
    chart.width = 18
    chart.height = 12

    data_ws = wb[data_ws_title]
    # Revenue is column 2 (index 1), header row included
    data_ref = Reference(data_ws, min_col=2, min_row=data_start_row, max_row=data_start_row + num_rows)
    cats_ref = Reference(data_ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"
    ws.add_chart(chart, chart_anchor)


def _add_line_chart(ws, data_ws_title: str, wb: Workbook, data_start_row: int, num_rows: int, chart_anchor: str, title: str):
    """Add a line chart to ws for monthly revenue trend."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    chart.width = 18
    chart.height = 12

    data_ws = wb[data_ws_title]
    data_ref = Reference(data_ws, min_col=2, min_row=data_start_row, max_row=data_start_row + num_rows)
    cats_ref = Reference(data_ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = "1F3864"
    chart.series[0].marker.symbol = "circle"
    ws.add_chart(chart, chart_anchor)


def build_report(
    df: pd.DataFrame,
    cat_summary: pd.DataFrame,
    monthly: pd.DataFrame,
    top_prods: pd.DataFrame,
    kpis: dict,
    output_path: str | Path,
):
    """Build and save the multi-sheet Excel report."""
    output_path = Path(output_path)
    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Executive Summary
    # ------------------------------------------------------------------ #
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"

    # Title
    ws_summary.merge_cells("A1:J1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Sales Performance Report"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    # KPI cards — row 3 (label) + row 4 (value), two columns each
    kpi_items = [
        ("Total Revenue", f"${kpis['total_revenue']:,.0f}"),
        ("Total Units", f"{kpis['total_units']:,}"),
        ("Avg Order Value", f"${kpis['avg_order_value']:,.0f}"),
        ("Top Product", kpis["top_product"]),
        ("Top Region", kpis["top_region"]),
    ]
    col_start = 1
    for label, value in kpi_items:
        col_end = col_start + 1
        # Label row
        ws_summary.merge_cells(
            start_row=3, start_column=col_start, end_row=3, end_column=col_end
        )
        lbl = ws_summary.cell(row=3, column=col_start, value=label)
        lbl.fill = KPI_LABEL_FILL
        lbl.font = KPI_LABEL_FONT
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        # Value row
        ws_summary.merge_cells(
            start_row=4, start_column=col_start, end_row=4, end_column=col_end
        )
        val = ws_summary.cell(row=4, column=col_start, value=value)
        val.fill = KPI_FILL
        val.font = KPI_VALUE_FONT
        val.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[3].height = 20
        ws_summary.row_dimensions[4].height = 36
        col_start += 2

    # ------------------------------------------------------------------ #
    # Sheet 2: By Category
    # ------------------------------------------------------------------ #
    ws_cat = wb.create_sheet("By Category")
    _write_df(ws_cat, cat_summary, start_row=1)

    # ------------------------------------------------------------------ #
    # Sheet 3: Monthly Trend
    # ------------------------------------------------------------------ #
    ws_monthly = wb.create_sheet("Monthly Trend")
    _write_df(ws_monthly, monthly, start_row=1)

    # Charts start at row 6 (added after data sheets exist)
    _add_bar_chart(ws_summary, "By Category", wb, 1, len(cat_summary), 2, "A6", "Revenue by Category")
    _add_line_chart(ws_summary, "Monthly Trend", wb, 1, len(monthly), "K6", "Monthly Revenue Trend")

    # ------------------------------------------------------------------ #
    # Sheet 4: Top Products
    # ------------------------------------------------------------------ #
    ws_top = wb.create_sheet("Top Products")
    _write_df(ws_top, top_prods, start_row=1)

    # ------------------------------------------------------------------ #
    # Sheet 5: Raw Data (drop _source)
    # ------------------------------------------------------------------ #
    ws_raw = wb.create_sheet("Raw Data")
    raw_df = df.drop(columns=["_source"], errors="ignore")
    _write_df(ws_raw, raw_df, start_row=1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
